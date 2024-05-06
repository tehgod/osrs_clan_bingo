import gspread
import json
import random
from gspread.cell import Cell
import string
from time import sleep
from openpyxl.utils import get_column_letter
from database import *

def open_json(filepath:str):
    with open (filepath) as myfile:
        return json.load(myfile)

class google_sheets:
    def __init__(self, sheets_json_data:dict, spreadsheet_name:str):

        self.gspread_client = gspread.service_account_from_dict(sheets_json_data)
        self.worksheet = self.gspread_client.open(spreadsheet_name)
        self.current_sheet = self.worksheet.get_worksheet(0)

    def a1_to_row_col(self, a1_format:str):
        row = []
        column = []
        for letter in a1_format:
            if letter.isnumeric():
                row.append(letter)
            else:
                column.append(letter)
        row = int("".join(row))
        column = "".join(column)
        num = 0
        for c in column:
            if c in string.ascii_letters:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        column = num
        return {"row":row, "column":column}
    
    def change_to_sheet(self, sheet_name:str, create_if_missing=False):
        if self.current_sheet.title != sheet_name:
            try:
                self.current_sheet = self.worksheet.worksheet(sheet_name)
            except gspread.exceptions.WorksheetNotFound:
                if create_if_missing == True:
                    self.current_sheet = self.worksheet.add_worksheet(title=sheet_name, rows=300, cols=100)

    def format_scoreboard_step_1(self, members_list:dict, board_template:list):
        #includes overall spreadsheet changes, including cell and column width, and merging cells.
        #determine max width and height of board
        furthest_column = 0
        lowest_row = 0
        for cell in board_template:
                cell_location = self.a1_to_row_col(cell["Cell"])
                if cell_location["row"]>lowest_row:
                    lowest_row=cell_location["row"]
                if cell_location["column"]>furthest_column:
                    furthest_column=cell_location["column"]
        last_member_row = 6+len(members_list)
        if last_member_row > lowest_row:
            lowest_row = last_member_row
        body = {
            "requests": [
                #update scoreboard columns width
                {
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": self.current_sheet.id,
                            "dimension": "COLUMNS",
                            "startIndex": 4,
                            "endIndex": 100
                        },
                        "properties": {
                            "pixelSize": 120
                        },
                        "fields": "pixelSize"
                    }
                },
                #update scoreboard rows height
                {
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": self.current_sheet.id,
                            "dimension": "ROWS",
                            "startIndex": 5,
                            "endIndex": 100
                        },
                        "properties": {
                            "pixelSize": 90
                        },
                        "fields": "pixelSize"
                    }
                },
                #update vertical black lines
                {
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": self.current_sheet.id,
                            "dimension": "COLUMNS",
                            "startIndex": 0,
                            "endIndex": 1
                        },
                        "properties": {
                            "pixelSize": 9
                        },
                        "fields": "pixelSize"
                    }
                },
                {
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": self.current_sheet.id,
                            "dimension": "COLUMNS",
                            "startIndex": 3,
                            "endIndex": 4
                        },
                        "properties": {
                            "pixelSize": 9
                        },
                        "fields": "pixelSize"
                    }
                },
                {
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": self.current_sheet.id,
                            "dimension": "COLUMNS",
                            "startIndex": furthest_column,
                            "endIndex": furthest_column+1
                        },
                        "properties": {
                            "pixelSize": 9
                        },
                        "fields": "pixelSize"
                    }
                },
                #update horizontal black lines
                {
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": self.current_sheet.id,
                            "dimension": "ROWS",
                            "startIndex": 0,
                            "endIndex": 1
                        },
                        "properties": {
                            "pixelSize": 9
                        },
                        "fields": "pixelSize"
                    }
                },
                {
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": self.current_sheet.id,
                            "dimension": "ROWS",
                            "startIndex": 4,
                            "endIndex": 5
                        },
                        "properties": {
                            "pixelSize": 9
                        },
                        "fields": "pixelSize"
                    }
                },
                {
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": self.current_sheet.id,
                            "dimension": "ROWS",
                            "startIndex": lowest_row,
                            "endIndex": lowest_row+1
                        },
                        "properties": {
                            "pixelSize": 9
                        },
                        "fields": "pixelSize"
                    }
                },
                #merge and center top left team name
                {
                    "mergeCells": {
                        "mergeType": "MERGE_ALL",
                        "range": {
                            "sheetId": self.current_sheet.id,
                            "startRowIndex": 1,
                            "endRowIndex": 4,
                            "startColumnIndex": 1,
                            "endColumnIndex": 3
                        }
                    }
                }
            ]
        }
        self.worksheet.batch_update(body)

    def format_scoreboard_step_2(self, members_list:dict, board_template:list):
        #Includes any text and specific cell FORMATTING such as coloring,bolding, and text wrapping
        cell_formats = []
        #format all scoreboard cells and background colors
        base_cell_format ={
            "borders": {
                "top": {
                    "style": "SOLID",
                },
                "bottom": {
                    "style": "SOLID",
                },
                "left": {
                    "style": "SOLID",
                },
                "right": {
                    "style": "SOLID",
                }
            },
            "horizontalAlignment": "CENTER",
            "verticalAlignment": "MIDDLE",
            "wrapStrategy": "WRAP"
        }

        for cell in board_template:
            current_cell_format = base_cell_format.copy()
            match cell["Difficulty"]:
                case "Start":
                    current_cell_format["backgroundColor"]={
                        "red": 0.0,
                        "green": 1.0,
                        "blue": 0.0
                    }
                    current_cell_format["textFormat"]={
                        "bold": True,
                        "fontSize": 20
                    }
                case "Beginner":
                    current_cell_format["backgroundColor"]={
                        "red": 204/255,
                        "green": 204/255,
                        "blue": 204/255
                    }
                case "Easy":
                    current_cell_format["backgroundColor"]={
                        "red": 217/255,
                        "green": 234/255,
                        "blue": 211/255
                    }
                case "Medium":
                    current_cell_format["backgroundColor"]={
                        "red": 207/255,
                        "green": 226/255,
                        "blue": 243/255
                    }
                case "Hard":
                    current_cell_format["backgroundColor"]={
                        "red": 234/255,
                        "green": 209/255,
                        "blue": 220/255
                    }
                case "Elite":
                    current_cell_format["backgroundColor"]={
                        "red": 252/255,
                        "green": 229/255,
                        "blue": 205/255
                    }
                case "Master":
                    current_cell_format["backgroundColor"]={
                        "red": 230/255,
                        "green": 184/255,
                        "blue": 175/255
                    }
            cell_formats.append(
                {
                    "range": cell["Cell"],
                    "format": current_cell_format
                }
            )
        #Bolden usernames and mvp points headers
        cell_formats.append(
            {
                "range": "B6:C6",
                "format": {
                    "textFormat":{
                        "bold": True,
                        "fontSize": 10
                    },
                    "backgroundColor": {
                        "red": 230/255,
                        "green": 145/255,
                        "blue": 56/255
                    },
                    "horizontalAlignment": "CENTER",
                    "verticalAlignment": "MIDDLE",
                    "wrapStrategy": "WRAP"
                }
            }
        )
        #bolden tasks and points values headers
        cell_formats.append(
            {
                "range": "E2:E4",
                "format": {
                    "textFormat":{
                        "bold": True,
                        "fontSize": 10
                    },
                    "backgroundColor": {
                        "red": 230/255,
                        "green": 145/255,
                        "blue": 56/255
                    },
                    "horizontalAlignment": "CENTER",
                    "verticalAlignment": "MIDDLE",
                    "wrapStrategy": "WRAP"
                }
            }
        )
        #bolden team name
        cell_formats.append(
            {
                "range": "B2:C4",
                "format": {
                    "textFormat":{
                        "bold": True,
                        "fontSize": 28
                    },
                    "backgroundColor": {
                        "red": 52/255,
                        "green": 168/255,
                        "blue": 83/255
                    },
                    "horizontalAlignment": "CENTER",
                    "verticalAlignment": "MIDDLE",
                    "wrapStrategy": "WRAP"
                }
            }
        )
        #determine max width and height of board
        furthest_column = 0
        lowest_row = 0
        for cell in board_template:
            cell_location = self.a1_to_row_col(cell["Cell"])
            if cell_location["row"]>lowest_row:
                lowest_row=cell_location["row"]
            if cell_location["column"]>furthest_column:
                furthest_column=cell_location["column"]
        last_member_row = 6+len(members_list)
        if last_member_row >lowest_row:
            lowest_row=last_member_row
        furthest_column+=1
        lowest_row+=1
        furthest_column_letter = get_column_letter(furthest_column)
        #fill in black veritcal bars
        ranges = [
            f"a1:{furthest_column_letter}1",
            f"a5:{furthest_column_letter}5",
            f"a{lowest_row}:{furthest_column_letter}{lowest_row}",
            f"a1:a{lowest_row}", 
            f"d1:d{lowest_row}", 
            f"{furthest_column_letter}1:{furthest_column_letter}{lowest_row}"
        ]
        for cell_range in ranges:
            cell_formats.append(
            {
                "range": cell_range,
                "format": {
                    "backgroundColor":{
                        "red": 0,
                        "green":0,
                        "blue": 0
                    },
                    "horizontalAlignment": "CENTER",
                    "verticalAlignment": "MIDDLE",
                    "wrapStrategy": "WRAP"
                }
            }
        )
        #set headers and other value colors in value board
        ranges = {
            "names":f"b7:c{str(last_member_row)}",
            "beginner":"f2:f4",
            "easy":"g2:g4",
            "medium":"h2:h4",
            "hard":"i2:i4",
            "elite":"j2:j4",
            "master":"k2:k4",
            "finished":"l2:l4"
        }
        for cell_range in ranges:
            base_cell_format ={
                "borders": {
                    "top": {
                        "style": "SOLID",
                    },
                    "bottom": {
                        "style": "SOLID",
                    },
                    "left": {
                        "style": "SOLID",
                    },
                    "right": {
                        "style": "SOLID",
                    }
                },
                "horizontalAlignment": "CENTER",
                "verticalAlignment": "MIDDLE",
                "wrapStrategy": "WRAP"
            }
            match cell_range:
                case "names":
                    base_cell_format["borders"]={
                        "top": {
                            "style": "SOLID",
                        },
                        "bottom": {
                            "style": "SOLID",
                        },
                        "left": {
                            "style": "NONE",
                        },
                        "right": {
                            "style": "NONE",
                        }
                    }
                    base_cell_format["backgroundColor"]={
                        "red": 246/255,
                        "green": 178/255,
                        "blue": 107/255
                    }
                case "beginner":
                    base_cell_format["backgroundColor"]={
                        "red": 204/255,
                        "green": 204/255,
                        "blue": 204/255
                    }
                case "easy":
                    base_cell_format["backgroundColor"]={
                        "red": 217/255,
                        "green": 234/255,
                        "blue": 211/255
                    }
                case "medium":
                    base_cell_format["backgroundColor"]={
                        "red": 207/255,
                        "green": 226/255,
                        "blue": 243/255
                    }
                case "hard":
                    base_cell_format["backgroundColor"]={
                        "red": 234/255,
                        "green": 209/255,
                        "blue": 220/255
                    }
                case "elite":
                    base_cell_format["backgroundColor"]={
                        "red": 252/255,
                        "green": 229/255,
                        "blue": 205/255
                    }
                case "master":
                    base_cell_format["backgroundColor"]={
                        "red": 230/255,
                        "green": 184/255,
                        "blue": 175/255
                    }
                case "finished":
                    base_cell_format["backgroundColor"]={
                        "red": 0/255,
                        "green": 255/255,
                        "blue": 0/255
                    }
            cell_formats.append(
                {
                    "range": ranges[cell_range],
                    "format": base_cell_format
                }
            )
        # #Push all changes
        self.current_sheet.batch_format(cell_formats)
        pass

    def format_scoreboard_step_3(self, members_list:dict, board_template:dict):
        cell_updates = []
        #add text for members, mvp points, and respective headers
        cell_updates.append(Cell(6, 2, "Members:"))
        cell_updates.append(Cell(6, 3, "MVP Points:"))
        current_row = 6
        for member in members_list:
            current_row+=1
            cell_updates.append(Cell(current_row, 2, member["Username"]))
        #add text for tiles, points, and respective headers
        cell_updates.append(Cell(2, 5, "Task"))
        cell_updates.append(Cell(2, 6, "Beginner"))
        cell_updates.append(Cell(2, 7, "Easy"))
        cell_updates.append(Cell(2, 8, "Medium"))
        cell_updates.append(Cell(2, 9, "Hard"))
        cell_updates.append(Cell(2, 10, "Elite"))
        cell_updates.append(Cell(2, 11, "Master"))
        cell_updates.append(Cell(2, 12, "Finished"))
        cell_updates.append(Cell(3, 5, "Points"))
        cell_updates.append(Cell(3, 6, 1))
        cell_updates.append(Cell(3, 7, 2))
        cell_updates.append(Cell(3, 8, 3))
        cell_updates.append(Cell(3, 9, 5))
        cell_updates.append(Cell(3, 10, 8))
        cell_updates.append(Cell(3, 11, 12))
        cell_updates.append(Cell(4, 5, "Total Possible"))
        #add text for team name
        cell_updates.append(Cell(2, 2, self.worksheet.title))
        #add text for start tile
        for tile in board_template:
            if tile["Difficulty"]=="Start":
                tile_location = self.a1_to_row_col(tile["Cell"])
                cell_updates.append(Cell(tile_location["row"], tile_location["column"], "START"))
        self.current_sheet.update_cells(cell_updates)
        pass

    def write_to_scoreboard(self, board_layout:list, individual_tile=None):
        if self.current_sheet.title != "Scoreboard":
            try:
                self.change_to_sheet("Scoreboard")
            except gspread.exceptions.WorksheetNotFound:
                print("No scoreboard located to write value to.")
                return False
        cells = []
        tiles_to_update = []
        if individual_tile==None:
            tiles_to_update = [tile["Cell"] for tile in board_layout]
        elif (individual_tile in [tile["Cell"] for tile in board_layout]) and type(individual_tile)==str:
            tiles_to_update.append(individual_tile)
        elif type(individual_tile)==list:
            tiles_to_update=individual_tile
        else:
            print("Inavlid tile selected")
            return False
        for tile_information in board_layout:
            if tile_information["Cell"] in tiles_to_update:
                cell_location = self.a1_to_row_col(tile_information["Cell"])
                cells.append(Cell(cell_location["row"], cell_location["column"], tile_information["Task"]))
        self.current_sheet.update_cells(cells)
        print("Successfully updated tiles.")
        return True

    def complete_tile(self, board_layout:list, individual_tile:str):
        if self.current_sheet.title != "Scoreboard":
            try:
                self.change_to_sheet("Scoreboard")
            except gspread.exceptions.WorksheetNotFound:
                print("No scoreboard located to write value to.")
                return False
        cell_update = {
            "range": individual_tile,
            "format": {
                "backgroundColor": {
                    "red": 0/255,
                    "green": 255/255,
                    "blue": 0/255
                }
            }
        }
        self.current_sheet.batch_format([cell_update])
        original_tile = self.a1_to_row_col(individual_tile)
        surrounding_tiles=[
            f"{get_column_letter(original_tile['column']-1)}{original_tile['row']}".lower(),
            f"{get_column_letter(original_tile['column']+1)}{original_tile['row']}".lower(),
            f"{get_column_letter(original_tile['column'])}{original_tile['row']-1}".lower(),
            f"{get_column_letter(original_tile['column'])}{original_tile['row']+1}".lower()
        ]
        valid_surrounding_tiles = []
        for tile in surrounding_tiles:
            if tile in [tile["Cell"] for tile in board_layout]:
                valid_surrounding_tiles.append(tile)
        self.write_to_scoreboard(board_layout, valid_surrounding_tiles)            

    def create_scoreboard(self, members_list:dict, board_template:dict):
        self.change_to_sheet("Scoreboard", True)
        self.format_scoreboard_step_1(members_list, board_template)
        self.format_scoreboard_step_2(members_list, board_template)
        self.format_scoreboard_step_3(members_list, board_template)

    def create_rules(self):
        pass

if __name__ == "__main__":
    my_db = database_connection()
    member_list = my_db.load_team_members(1)
    board_layout = my_db.load_board_layout(1)
    board_template = my_db.load_template(1)
    team_1_creds = my_db.load_team_credentials(1)
    team_1_sheets = google_sheets(json.loads(team_1_creds[0]["GoogleSheetAuth"]), team_1_creds[0]["Name"])

    team_1_sheets.create_scoreboard(member_list, board_template)
    team_1_sheets.complete_tile(board_layout, "i10")
    team_1_sheets.write_to_scoreboard(board_layout, "i13")
